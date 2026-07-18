#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
录播整理文档增量更新脚本 (纯 Python 版)
==========================================================================
作用：读取项目库(bili-clip-flow.sqlite3)里的投稿任务(所有账号) + 百度网盘路径，
      把“上次更新之后新建的投稿任务”按现有《主播-分类》结构追加进 xlsx。

增量机制(基于时间水位)：
  * 文档里维护一个《更新记录》sheet，记录每次运行的时间水位。
  * 水位 = 运行时“库中投稿任务的最新创建时间(MAX created_at)”，格式 yyyy-MM-dd HH:mm:ss。
  * 每次运行读取上次水位，只查询 created_at >= 上次水位 的任务(重扫那一秒),
    再靠 BV 去重防重复；跑完把水位推进到最新。
  * 首次运行(无《更新记录》)：扫描全部→BV 去重后一般 0 新增→写入基线水位。

分类规则(与人工整理一致)：
  ① 该作品已在文档中出现 → 沿用它所在分类(自描述)
  ② 网盘路径含分类目录(anime/动漫/movie/电影/电视剧/影视/特摄…) → 自动归类
  ③ 都没有 → 放进《主播-未分类》表，待手动移动(下次自动沿用)
  视频名称取路径末级文件夹(季/剧场版/part 自动上取一级)；同名多集并组、按集数排序、合并单元格。
  无网盘路径任务、SKIP_BVS 里已人工删除的任务 跳过。

待确认(黄色)机制：
  * 每次更新新增的视频行会被填成【黄色】，供你核对。黄色会跨多次运行保留(重建时自动重上色)。
  * 核对无误后，用 --confirm 清除全部黄色标记(不新增任务)。

用法：
  python3 update_reaction_doc.py                 # 增量更新(新增行标黄, 先自动备份)
  python3 update_reaction_doc.py --dry-run       # 只预览，不写文件
  python3 update_reaction_doc.py --confirm       # 确认完成: 清除全部黄色标记
  python3 update_reaction_doc.py --init          # 文档不存在时新建空白并从库全量生成
  python3 update_reaction_doc.py <db路径> <xlsx路径>

依赖：python3 + openpyxl (pip3 install openpyxl)   # 数据库用 python 自带 sqlite3
==========================================================================
"""
import argparse, os, re, sqlite3, shutil, datetime, sys

# ============ 可调配置 ============
LOG_SHEET = "更新记录"
LOG_HDR   = ["运行时间(本地)", "上次更新水位", "本次更新水位(最新任务创建时间)", "本次新增行数"]
CATORDER  = ["番剧", "电影", "电视剧", "特摄", "未分类"]
PATHCAT   = {"anime": "番剧", "动漫": "番剧", "动画": "番剧", "番剧": "番剧",
             "movie": "电影", "电影": "电影",
             "电视剧": "电视剧", "影视": "电视剧", "tv": "电视剧", "剧集": "电视剧",
             "特摄": "特摄", "tokusatsu": "特摄"}
SKIP_BVS  = {"BV1CJZ8BPE5Y"}   # 已人工删除、永不再加(曜阳川王杯乐龄春晚)
ACC       = {82679456: "明前奶姊(82679456)", 3546735668366092: "明前奶姊的小号(3546735668366092)"}
HDR       = ["视频名称", "网盘路径", "对应BV号", "对应账号"]
WIDTHS    = [36, 150, 37, 28]
# =================================

GENERIC = re.compile(r"^(season\s*\d*|s\d+|part\s*\d*|ova|oad|sp|第[一二三四五六七八九十\d]+[季部]|剧场版.*|\d+~\d+)$", re.I)


def streamer_of(p):
    if not p:
        return None
    p = p if p.startswith("/") else "/" + p
    m = re.match(r"^/Bilibili/Reaction/([^/]+)", p)
    return m.group(1) if m else None


def video_name(p):
    parts = [x for x in p.rstrip("/").split("/") if x]
    body = parts[3:] if len(parts) > 3 else []
    if not body:
        return None
    name = body[-1]
    if GENERIC.match(name.strip()) and len(body) >= 2:
        name = body[-2]
    return name


def path_cat(p):
    if not p:
        return None
    parts = [x for x in p.strip("/").split("/") if x]
    mids = parts[3:-1] if len(parts) > 4 else (parts[3:] if len(parts) > 3 else [])
    for f in mids:
        if f.strip().lower() in PATHCAT:
            return PATHCAT[f.strip().lower()]
        if f.strip() in PATHCAT:
            return PATHCAT[f.strip()]
    return None


def full_path(p, f):
    return p.rstrip("/") + "/" + str(f).strip() if (f and str(f).strip()) else p


def fmt_ts(v):
    # 把 ISO(2026-07-04T12:58:53.660049+00:00) 或已格式化值 归一为 yyyy-MM-dd HH:mm:ss
    if v is None:
        return None
    return str(v).strip().replace("T", " ")[:19]


def natkey(row):
    fn = str(row[0]).rstrip("/").split("/")[-1]
    fn = re.sub(r"\.(mp4|mkv|flv|mov|ts)$", "", fn, flags=re.I)
    m = re.search(r"(\d+)\s*~\s*\d+", fn) or re.search(r"\d+", fn)
    return (int(m.group(1) if m and m.lastindex else (m.group() if m else 10**9)) if m else 10**9, fn)


def sheet_cat(title):
    if "-" in title:
        s, c = title.rsplit("-", 1)
        if c in CATORDER:
            return s, c
    return title, None


# ---------- 数据库自动发现 ----------
DB_FILENAME = "bili-clip-flow.sqlite3"


def _has_submission_table(path):
    """校验该 sqlite 文件确实是本项目库(含 submission_task 表)。只读打开, 不加锁不创建。"""
    try:
        con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        ok = con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='submission_task'"
        ).fetchone() is not None
        con.close()
        return ok
    except Exception:
        return False


def _candidate_roots():
    """各系统里 app 数据可能所在的根目录。"""
    home = os.path.expanduser("~")
    roots = [
        os.path.join(home, "Library", "Application Support"),   # macOS
        os.path.join(home, ".local", "share"),                  # Linux
        os.path.join(home, ".config"),                          # Linux
        os.environ.get("APPDATA", ""),                          # Windows
        os.environ.get("LOCALAPPDATA", ""),                     # Windows
        os.path.dirname(os.path.abspath(__file__)),             # 脚本同目录
        os.getcwd(),                                            # 当前目录
    ]
    seen = []
    for r in roots:
        if r and os.path.isdir(r) and r not in seen:
            seen.append(r)
    return seen


def _walk_find(root, match, max_depth=6):
    """在 root 下按文件名匹配递归查找(限制深度), 返回所有命中路径。"""
    hits = []
    for dirpath, dirnames, filenames in os.walk(root):
        if dirpath[len(root):].count(os.sep) >= max_depth:
            dirnames[:] = []
            continue
        for fn in filenames:
            if match(fn):
                hits.append(os.path.join(dirpath, fn))
    return hits


def discover_db(preferred):
    """按 首选路径 → 标准子路径 → 目录搜文件名 → 搜含submission_task的库 顺序发现数据库。
    返回 (路径 or None, 来源标记)。"""
    home = os.path.expanduser("~")
    # 1) 首选(默认/传入)路径直接可用
    if preferred and os.path.isfile(preferred):
        return preferred, "default"
    # 2) 各系统标准子路径
    for e in [
        os.path.join(home, "Library", "Application Support", "com.tbw.biliclipflow", DB_FILENAME),
        os.path.join(os.environ.get("APPDATA", ""), "com.tbw.biliclipflow", DB_FILENAME),
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "com.tbw.biliclipflow", DB_FILENAME),
    ]:
        if e and os.path.isfile(e):
            return e, "standard"
    # 3) 候选根目录下按精确文件名搜, 优先含 submission_task 表的
    for root in _candidate_roots():
        hits = _walk_find(root, lambda fn: fn == DB_FILENAME)
        valid = [h for h in hits if _has_submission_table(h)]
        if valid:
            return valid[0], "search"
        if hits:
            return hits[0], "search"
    # 4) 兜底: 搜任意含 submission_task 表的 .sqlite3/.db
    for root in _candidate_roots():
        for h in _walk_find(root, lambda fn: fn.endswith((".sqlite3", ".db"))):
            if _has_submission_table(h):
                return h, "search-table"
    return None, None


def run(DB, XLSX, DRY, CONFIRM=False):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    F = Font(name="宋体", size=11)
    CEN = Alignment(vertical="center")
    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def is_yellow(cell):
        fl = cell.fill
        return bool(fl) and fl.patternType == "solid" and str(getattr(fl.fgColor, "rgb", "")).upper().endswith("FFFF00")

    pending = set()   # (网盘路径, BV) —— 待确认(黄色)的行

    # ---------- 1) 解析当前文档 ----------
    wb = openpyxl.load_workbook(XLSX)
    model = {}; doc_bvs = set(); work_cat = {}; streamer_order = []
    log_history = []; last_watermark = None
    for title in wb.sheetnames:
        if title == LOG_SHEET:
            lws = wb[title]
            for r in lws.iter_rows(min_row=2, values_only=True):
                r = list(r) + [None] * 4
                if r[0] is None and r[2] is None:
                    continue
                log_history.append([r[0], r[1], r[2], r[3]])
                if r[2]:
                    last_watermark = fmt_ts(r[2])   # 取最后一行“本次水位”(归一化, 兼容旧ISO值)
            continue
        s, cat = sheet_cat(title)
        if cat is None:                              # 其它非结构表：原样保留
            continue
        if s not in streamer_order:
            streamer_order.append(s)
        ws = wb[title]; am = {}
        for m in ws.merged_cells.ranges:
            ms = str(m)
            if ms.startswith("A"):
                am[int(ms.split(":A")[0][1:])] = int(ms.split(":A")[1])
        C = lambda r, c: ws.cell(row=r, column=c).value
        i = 2; mx = ws.max_row
        while i <= mx:
            bot = am[i] if i in am else i
            rng = range(i, bot + 1)
            rws = [(C(r, 2), C(r, 3), C(r, 4)) for r in rng]
            nm = C(i, 1); mg = i in am
            for r in rng:                                    # 记住现有的黄色(待确认)行, 重建时保留
                if is_yellow(ws.cell(r, 2)):
                    pending.add((C(r, 2), C(r, 3)))
            model.setdefault((s, cat), []).append({"name": nm, "rows": rws, "merged": mg, "touched": False})
            if nm is not None:
                work_cat.setdefault((s, str(nm)), cat)
            for _, bv, _ in rws:
                if bv and str(bv).startswith("BV"):
                    doc_bvs.add(str(bv))
            i = bot + 1

    # ---------- 2) 按时间水位查询新任务 ----------
    con = sqlite3.connect(DB); con.row_factory = sqlite3.Row
    NORM = "substr(replace(created_at,'T',' '),1,19)"   # created_at 归一为 yyyy-MM-dd HH:mm:ss 再比较
    # 用 >= 重扫“水位那一秒”，靠 BV 去重防重复：可捕获与水位同秒、以及基线时被跳过的任务，避免漏加。
    if last_watermark:
        rows = con.execute("SELECT bvid,bilibili_uid,baidu_sync_path,baidu_sync_filename,created_at "
                           f"FROM submission_task WHERE bvid IS NOT NULL AND bvid!='' AND {NORM} >= ? "
                           "ORDER BY created_at", (last_watermark,)).fetchall()
    else:
        rows = con.execute("SELECT bvid,bilibili_uid,baidu_sync_path,baidu_sync_filename,created_at "
                           "FROM submission_task WHERE bvid IS NOT NULL AND bvid!='' ORDER BY created_at").fetchall()
    new_watermark = fmt_ts(con.execute("SELECT MAX(created_at) FROM submission_task").fetchone()[0])

    if CONFIRM:                 # 确认完成模式: 不新增任务, 清除全部待确认(黄色)
        rows = []
        cleared = len(pending)
        pending.clear()

    added = 0; skip_nopath = 0; per_sheet = {}
    for r in rows:
        bv = r["bvid"]
        if bv in doc_bvs or bv in SKIP_BVS:
            continue
        p = r["baidu_sync_path"]
        if not p:
            skip_nopath += 1; continue
        s = streamer_of(p)
        if not s:
            skip_nopath += 1; continue
        name = video_name(p); path = full_path(p, r["baidu_sync_filename"])
        acc = ACC.get(r["bilibili_uid"], str(r["bilibili_uid"]))
        cat = work_cat.get((s, str(name))) or path_cat(p) or "未分类"
        if s not in streamer_order:
            streamer_order.append(s)
        units = model.setdefault((s, cat), [])
        hit = next((u for u in units if u["name"] is not None and str(u["name"]) == str(name)), None)
        if hit is None:
            hit = {"name": name, "rows": [], "merged": False, "touched": True}; units.append(hit)
        hit["rows"].append((path, bv, acc)); hit["touched"] = True
        pending.add((path, bv))                # 新增行标黄, 待你确认
        doc_bvs.add(bv); added += 1
        per_sheet[f"{s}-{cat}"] = per_sheet.get(f"{s}-{cat}", 0) + 1

    for units in model.values():
        for u in units:
            if u["touched"] and len(u["rows"]) > 1:
                u["rows"].sort(key=natkey)

    # ---------- 3) 重建工作簿 ----------
    def write_units(ws, units):
        for ci, w in zip("ABCD", WIDTHS):
            ws.column_dimensions[ci].width = w
        for ci, h in enumerate(HDR, 1):
            ws.cell(row=1, column=ci, value=h).font = F
        rrow = 2; merges = []
        for u in units:
            start = rrow
            for k, (b, c, d) in enumerate(u["rows"]):
                mark = (b, c) in pending                     # 待确认 → 黄色底
                for col, val in ((1, u["name"] if k == 0 else None), (2, b), (3, c), (4, d)):
                    cell = ws.cell(row=rrow, column=col, value=val)
                    cell.font = F
                    if mark:
                        cell.fill = YELLOW
                rrow += 1
            if len(u["rows"]) > 1 and (u["merged"] or u["name"] is not None):
                merges.append((start, rrow - 1))
        for a, b in merges:
            ws.merge_cells(f"A{a}:A{b}"); ws.cell(row=a, column=1).alignment = CEN

    now_local = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _has_data(ws):   # 是否有任何非空单元格(用于剔除空白/占位 sheet)
        return any(any(v not in (None, "") for v in row) for row in ws.iter_rows(values_only=True))

    keep_other = [t for t in wb.sheetnames
                  if t != LOG_SHEET and sheet_cat(t)[1] is None and _has_data(wb[t])]
    newwb = openpyxl.Workbook(); newwb.remove(newwb.active)

    # 3a) 《更新记录》放在最前
    log = newwb.create_sheet(title=LOG_SHEET)
    log.column_dimensions["A"].width = 20; log.column_dimensions["B"].width = 34
    log.column_dimensions["C"].width = 34; log.column_dimensions["D"].width = 12
    for ci, h in enumerate(LOG_HDR, 1):
        log.cell(row=1, column=ci, value=h).font = F
    lr = 2
    for row in log_history:
        for ci, v in enumerate(row, 1):
            log.cell(row=lr, column=ci, value=v).font = F
        lr += 1
    if not DRY and not CONFIRM:   # 仅真实“更新”时追加记录; 确认模式不算一次更新
        for ci, v in enumerate([now_local, last_watermark or "-", new_watermark, added], 1):
            log.cell(row=lr, column=ci, value=v).font = F

    # 3b) 其它非结构表原样保留
    for t in keep_other:
        src = wb[t]; dst = newwb.create_sheet(title=t)
        for row in src.iter_rows(values_only=True):
            dst.append(list(row))

    # 3c) 主播 × 分类
    total = 0
    for s in streamer_order:
        for cat in CATORDER:
            units = model.get((s, cat))
            if not units:
                continue
            write_units(newwb.create_sheet(title=f"{s}-{cat}"[:31]), units)
            total += sum(len(u["rows"]) for u in units)

    # ---------- 4) 输出 ----------
    if CONFIRM:
        print(f"\n确认完成模式: 清除 {cleared} 行待确认(黄色)标记。")
    else:
        print(f"\n上次水位: {last_watermark or '(无, 首次)'}")
        print(f"本次水位: {new_watermark}")
        print(f"新增行:   {added}    跳过(无网盘路径): {skip_nopath}")
        if per_sheet:
            print("新增分布:")
            for k, v in sorted(per_sheet.items()):
                print(f"   + {k}: {v}")
    print(f"结果:     {len(newwb.sheetnames)} 个 sheet, {total} 数据行, 待确认(黄色)行 {len(pending)}")

    if DRY:
        print("\n[DRY-RUN] 未写入文件。")
    else:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = XLSX[:-5] + f".bak.{ts}.xlsx" if XLSX.endswith(".xlsx") else XLSX + f".bak.{ts}"
        shutil.copy2(XLSX, bak)
        newwb.save(XLSX)
        if CONFIRM:
            print(f"\n✅ 已确认: 清除全部黄色标记 -> {XLSX}")
        else:
            print(f"\n✅ 已更新: {XLSX}")
            print(f"   水位已推进至: {new_watermark}")
        print(f"   备份: {bak}")
        if added or CONFIRM:
            print("   记得重新上传到腾讯文档。")


def main():
    home = os.path.expanduser("~")
    default_db = os.path.join(home, "Library", "Application Support", "com.tbw.biliclipflow", "bili-clip-flow.sqlite3")
    default_xlsx = os.path.join(home, "Downloads", "录播整理2.xlsx")

    ap = argparse.ArgumentParser(description="录播整理文档增量更新脚本")
    ap.add_argument("db", nargs="?", default=None, help="数据库路径 (默认: 自动发现 app 数据库)")
    ap.add_argument("xlsx", nargs="?", default=None, help="xlsx 文档路径 (默认: ~/Downloads/录播整理2.xlsx)")
    ap.add_argument("--dry-run", "-n", action="store_true", help="只预览, 不写文件")
    ap.add_argument("--confirm", action="store_true", help="确认完成: 清除全部待确认(黄色)标记, 不新增任务")
    ap.add_argument("--init", action="store_true", help="文档不存在时新建空白文档并从库全量生成(无分类目录的作品会进未分类, 需人工归)")
    args = ap.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("❌ 缺少 openpyxl，请先执行: pip3 install openpyxl")
        sys.exit(1)

    # 数据库：显式传参优先；否则用默认路径，默认不存在则自动发现
    if args.db:
        db = args.db
        if not os.path.isfile(db):
            print(f"❌ 指定的数据库不存在: {db}")
            sys.exit(1)
    else:
        db, how = discover_db(default_db)
        if not db:
            print("❌ 未找到数据库 bili-clip-flow.sqlite3。")
            print("   请手动指定: python3 update_reaction_doc.py <数据库路径> <xlsx路径>")
            sys.exit(1)
        if how != "default":
            print(f"🔎 自动发现数据库: {db}")

    xlsx = args.xlsx or default_xlsx
    if not os.path.isfile(xlsx):
        if args.init:
            openpyxl.Workbook().save(xlsx)
            print(f"🆕 文档不存在，已新建空白文档: {xlsx}")
        else:
            print(f"❌ 文档不存在: {xlsx}")
            print("   若要从零生成整个文档，请加 --init")
            sys.exit(1)

    print(f"数据库: {db}")
    print(f"文档:   {xlsx}")
    if args.dry_run:
        print("模式:   DRY-RUN (仅预览)")
    if args.confirm:
        print("模式:   确认完成 (清除黄色标记)")

    run(db, xlsx, args.dry_run, args.confirm)


if __name__ == "__main__":
    main()
